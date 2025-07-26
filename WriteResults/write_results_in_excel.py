import pandas as pd
import os
import re
from openpyxl import load_workbook

# def write_data_excel(data, filename):
#     df = pd.DataFrame(data) #Make data in a Pandas DataFrame 
#     df.to_excel(filename, index=False, engine="openpyxl")
#     print(f"The results in the file: {filename}")
def write_data_excel(data, filename):
    try:
        # Try to load existing workbook
        book = load_workbook(filename)
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Convert data to DataFrame and write to Excel
            # Determine the next available row
            if 'Sheet1' in book.sheetnames:
                start_row = book['Sheet1'].max_row
            else:
                start_row = 0
            
            # Convert data to DataFrame and write to Excel
            if isinstance(data, pd.DataFrame):
                data.to_excel(writer, index=False, header=start_row == 0, startrow=start_row, sheet_name='Sheet1')
            elif isinstance(data, dict):
                # df = pd.DataFrame(data)
                df = pd.DataFrame({key: [value] for key, value in data.items()})
                df.to_excel(writer, index=False, header=start_row == 0, startrow=start_row, sheet_name='Sheet1')
            else:
                df = pd.DataFrame(data, index=[0])  # Wrap data in a list to create a single-row DataFrame
                df.to_excel(writer, index=False, header=start_row == 0, startrow=start_row, sheet_name='Sheet1')
        print(f"Data appended to the file: {filename}")
    except FileNotFoundError:
        # If file doesn't exist, create a new one
        if isinstance(data, pd.DataFrame):
            print("df is a pandas DataFrame")
            data.to_excel(filename, index=False, engine="openpyxl")
        else:
            print("df is not a pandas DataFrame")
            df = pd.DataFrame([data])
            df.to_excel(filename, index=False, engine="openpyxl")
        print(f"New file created: {filename}")

def sanitize_filename(filename):
    # Remove invalid characters
    filename = re.sub(r'[<>:"/\\|?*]', '', filename)
    
    # Replace spaces with underscores
    filename = filename.replace(' ', '_')
    
    # Remove any non-ASCII characters
    filename = ''.join(char for char in filename if ord(char) < 128)
    
    # Truncate filename if it's too long (adjust max_length as needed)
    max_length = 255  # Maximum filename length for most file systems
    if len(filename) > max_length:
        name, ext = os.path.splitext(filename)
        filename = name[:max_length - len(ext)] + ext
    
    return filename